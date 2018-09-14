import requests
from bs4 import BeautifulSoup
from user_agent import generate_user_agent
import re
import openpyxl
from time import sleep


#задаем прокси и юзер агент

headers = {
    'User-Agent': generate_user_agent(device_type="desktop",os=('mac','linux'))
}

proxies = {'http' : 'http://10.10.0.0:0000',
            'https': 'http://120.10.0.0:0000'
           }


show_title = []
show_date = []
show_title_eng = []
show_counry = []
show_time = []
show_channel = []
show_watch = []
show_all_time = []
show_imdb = []
show_kinopoisk = []
show_MyShows = []
show_janr= []


#получаем список сериалов, которые смотрю
url = 'https://myshows.me/АнтонХохлов'
r = requests.get(url, headers=headers)
#r = requests.get(url, proxies=proxies)

soup = BeautifulSoup(r.text, "html.parser")

show_list = soup.find('table',{'class':'catalogTable _progress'})
shows = show_list.select('tr a')

links = []
for show in shows:
   links.append(show.get('href'))


links.remove('#') #удалить # из списка списке


#получаем информацию о сериалах
for link in links:
    r=requests.get(link,headers=headers)

    soup=BeautifulSoup(r.text,"html.parser")

    #Получаем заголовок
    show_title.append(soup.select("h1")[0].text)
    show_title_eng.append(soup.select(".subHeader")[0].text)

    info = soup.select(".clear")[0].text
    #print(info)

    show_date.append(re.search('Даты выхода: (.*)',info)[1])
    show_counry.append(re.search('Страна: (.*)', info)[1])
    show_time.append(re.search('серии:(.*). Рейтинг IMDB', info)[1])
    show_channel.append(re.search('Канал: (.*) Смотрящих:',info)[1])
    show_watch.append(re.search('Смотрящих: (\d{1,3}(\s\d{1,3})?)',info)[1])
    show_all_time.append(re.search('Общая длительность: (.*) Дли',info)[1])
    show_imdb.append(re.search('IMDB: (\d{1,2}(\.\d{1,3})?)',info)[1])
    #show_imdb.append(re.search('IMDB: (.*) из ', info)[1])
    show_kinopoisk.append(re.search('Кинопоиска: (\d{1,2}(\.\d{1,3})?)',info)[1])
    show_MyShows.append(re.search('MyShows: (\d{1,2}(\.\d{1,3})?)',info)[1])
    show_janr_bad = re.search('Жанры:\\n(.*)Канал' ,info,re.DOTALL)[1]
    show_janr_notspace = show_janr_bad.replace(" ","")
    show_janr.append(show_janr_notspace.replace(",",", "))


# создаем файл
wb = openpyxl.load_workbook(filename='/Users/toxa7333/Desktop/MyShow.xlsx')
sheet = wb['MyShow']

# сохраняем в xls

i=2

for link in links:
    sheet.cell(row=i, column=1).value = show_title[i-2]
    sheet.cell(row=i, column=2).value = show_title_eng[i-2]
    sheet.cell(row=i, column=3).value = show_date[i-2]
    sheet.cell(row=i, column=4).value = show_counry[i-2]
    sheet.cell(row=i, column=5).value = show_janr[i-2]
    sheet.cell(row=i, column=6).value = show_channel[i-2]
    sheet.cell(row=i, column=7).value = show_watch[i-2]
    sheet.cell(row=i, column=8).value = show_all_time[i-2]
    sheet.cell(row=i, column=9).value = show_time[i-2]
    sheet.cell(row=i, column=10).value = show_imdb[i-2]
    sheet.cell(row=i, column=11).value = show_kinopoisk[i-2]
    sheet.cell(row=i, column=12).value = show_MyShows[i-2]
    i =i + 1


wb.save('/Users/toxa7333/Desktop/MyShow.xlsx')
